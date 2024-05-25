from DTOs.Databases.user_feedback_dtos import UserFeedbackDTO
from Interfaces.Interactor.record_user_feedback_interactor_interface import \
    RecordUserFeedBackInteractorInterface
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


class RecordUserFeedBackInteractor(RecordUserFeedBackInteractorInterface):
    def record_user_feedback(self, user_feedback_dto: UserFeedbackDTO) -> str:
        """
            Records user feedback in an Excel sheet named 'user_feedback_sheet.xlsx' located
            within a folder named 'User_Feedback' relative to the script's location.

            Args:
                user_feedback_dto (UserFeedbackDTO): A Data Transfer Object containing the user's feedback.

            Returns:
                str: A message indicating success or failure (e.g., permission error).

            Raises:
                (Optional) Exceptions that may occur during the process:
                    - FileNotFoundError: If the Excel file doesn't exist.
                    - PermissionError: If the file is open and cannot be written to.
            """
        file_path = './../User_Feedback/user_feedback_sheet.xlsx'

        user_feedback_dict = {
            'time': user_feedback_dto.time,
            'date': user_feedback_dto.date,
            'user_email': user_feedback_dto.user_email,
            'user_input': user_feedback_dto.user_input,
            'response_received': user_feedback_dto.response_received,
            'time_taken': user_feedback_dto.time_taken,
            'feedback': user_feedback_dto.feedback
        }
        try:

            try:
                wb = openpyxl.load_workbook(filename=file_path)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            sheet = wb.get_sheet_by_name("Sheet1")

            if sheet["A1"].value is None:
                headers = list(user_feedback_dict.keys())
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col).value = header
                    sheet.cell(row=1, column=col).alignment = Alignment(
                        horizontal="center")

            row = sheet.max_row + 1
            for col, value in enumerate(user_feedback_dict.values(), 1):
                sheet.cell(row=row, column=col).value = value

            wb.save(filename=file_path)

            return f"Data has been successfully written to {file_path}"
        except PermissionError:
            return f"Please close the file {file_path}"
